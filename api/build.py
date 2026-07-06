"""
Vercel Python serverless function: build the final .xlsx (step 2 of 2).
Self-contained (no cross-file imports) to avoid Vercel Python bundling issues.

Takes the header/rows the user already reviewed in the preview step,
plus the target mode/filename/sheet, and returns the finished spreadsheet.
"""

import re
import io
import base64
import logging
from flask import Flask, request, jsonify
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

app = Flask(__name__)
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger("pdf-build")

MAX_EXISTING_XLSX_BYTES = 15 * 1024 * 1024
XLSX_MAGIC = b"PK"


def autosize_columns(ws):
    for col_cells in ws.columns:
        max_len = max((len(str(c.value)) for c in col_cells if c.value is not None), default=10)
        col_letter = get_column_letter(col_cells[0].column)
        ws.column_dimensions[col_letter].width = min(max_len + 2, 60)


def build_new_workbook(header, rows, sheet_name):
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    ws.append(header)
    for row in rows:
        ws.append(row)
    autosize_columns(ws)
    return wb


def append_to_workbook(existing_bytes, header, rows, sheet_name):
    wb = load_workbook(io.BytesIO(existing_bytes))
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        if ws.max_row == 0 or ws.cell(row=1, column=1).value is None:
            ws.append(header)
    else:
        ws = wb.create_sheet(sheet_name)
        ws.append(header)
    for row in rows:
        ws.append(row)
    autosize_columns(ws)
    return wb


def sanitize_filename(name, fallback="extracted_data.xlsx"):
    if not isinstance(name, str) or not name.strip():
        return fallback
    name = name.strip().replace("/", "").replace("\\", "")
    name = re.sub(r"[^A-Za-z0-9 ._-]", "", name)
    name = name.strip(" .")
    if not name:
        return fallback
    if not name.lower().endswith(".xlsx"):
        name += ".xlsx"
    return name[:120]


def sanitize_sheet_name(name, fallback="Extracted"):
    if not isinstance(name, str) or not name.strip():
        return fallback
    name = re.sub(r"[\[\]:*?/\\]", "", name).strip()
    return (name or fallback)[:31]


@app.route("/api/build", methods=["POST"])
def build():
    data = request.get_json(silent=True)
    if not data:
        return jsonify({"error": "Invalid or missing request body."}), 400

    header = data.get("header")
    rows = data.get("rows")
    if not header or rows is None:
        return jsonify({"error": "Missing extracted data. Please re-run extraction."}), 400

    mode = data.get("mode", "new")
    sheet_name = sanitize_sheet_name(data.get("sheet_name"))
    filename = sanitize_filename(data.get("filename"))
    existing_b64 = data.get("existing_xlsx_base64")

    existing_bytes = None
    if mode == "append" and existing_b64:
        try:
            existing_bytes = base64.b64decode(existing_b64, validate=True)
        except Exception:
            return jsonify({"error": "The existing spreadsheet data was corrupted in transit."}), 400

        if len(existing_bytes) > MAX_EXISTING_XLSX_BYTES:
            return jsonify({"error": "The existing spreadsheet is too large."}), 400
        if not existing_bytes.startswith(XLSX_MAGIC):
            return jsonify({"error": "That existing file doesn't look like a valid .xlsx spreadsheet."}), 400

    try:
        if mode == "append" and existing_bytes:
            wb = append_to_workbook(existing_bytes, header, rows, sheet_name)
        else:
            wb = build_new_workbook(header, rows, sheet_name)
    except Exception:
        logger.exception("Workbook build failed")
        return jsonify({"error": "Couldn't build the spreadsheet. The existing file may be corrupted or in an unsupported format."}), 400

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    out_b64 = base64.b64encode(buf.read()).decode("utf-8")

    logger.info("Built workbook: mode=%s, sheet=%s, rows=%s, filename=%s", mode, sheet_name, len(rows), filename)

    return jsonify({
        "xlsx_base64": out_b64,
        "filename": filename,
        "row_count": len(rows),
    })
