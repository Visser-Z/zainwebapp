"""
Shared extraction/workbook helpers used by both api/extract.py (preview)
and api/build.py (final write). Not a Flask app itself — just a module.
"""

import re
import io
import pdfplumber
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

NUMERIC_PATTERN = re.compile(r"^-?\d{1,3}(,\d{3})*(\.\d+)?$|^-?\d+(\.\d+)?$")

MAX_PDF_BYTES = 15 * 1024 * 1024
MAX_EXISTING_XLSX_BYTES = 15 * 1024 * 1024
PDF_MAGIC = b"%PDF-"
XLSX_MAGIC = b"PK"


def to_number_if_possible(value):
    if not isinstance(value, str) or value == "":
        return value
    stripped = value.replace(",", "")
    if NUMERIC_PATTERN.match(value):
        try:
            if "." in stripped:
                return float(stripped)
            return int(stripped)
        except ValueError:
            return value
    return value


def _row_matches_header(clean_row, header_without_page):
    """True if a data row is actually a repeated header (common on multi-page tables)."""
    if len(clean_row) != len(header_without_page):
        return False
    return [c.strip().lower() for c in clean_row] == [h.strip().lower() for h in header_without_page]


def extract_rows_from_bytes(pdf_bytes):
    """Returns (header_or_None, rows, used_real_tables). header/rows exclude typing decisions
    beyond number conversion; 'page' is always column 0."""
    all_rows = []
    header = None
    header_without_page = None
    any_tables_found = False

    with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
        for page_num, page in enumerate(pdf.pages, start=1):
            tables = page.extract_tables()

            if tables:
                any_tables_found = True
                for table in tables:
                    if not table:
                        continue
                    start_idx = 0
                    if header is None:
                        first_row = ["" if c is None else str(c).strip() for c in table[0]]
                        header = ["page"] + first_row
                        header_without_page = first_row
                        start_idx = 1

                    for row in table[start_idx:]:
                        clean_row = ["" if cell is None else str(cell).strip() for cell in row]

                        # Skip rows that are just a repeated header (common when a table
                        # continues across a page break and the PDF repeats the header row)
                        if header_without_page and _row_matches_header(clean_row, header_without_page):
                            continue

                        typed_row = [to_number_if_possible(v) for v in clean_row]
                        all_rows.append([page_num] + typed_row)
            else:
                text = page.extract_text() or ""
                if not text.strip():
                    continue
                if header is None:
                    header = ["page", "text_line"]
                for line in text.split("\n"):
                    if line.strip():
                        all_rows.append([page_num, line.strip()])

    return header, all_rows, any_tables_found


def autosize_columns(ws):
    for col_cells in ws.columns:
        max_len = max((len(str(c.value)) for c in col_cells if c.value is not None), default=10)
        col_letter = get_column_letter(col_cells[0].column)
        ws.column_dimensions[col_letter].width = min(max_len + 2, 60)


def build_new_workbook(header, rows, sheet_name):
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    ws.append(header)
    for row in rows:
        ws.append(row)
    autosize_columns(ws)
    return wb


def get_existing_row_signatures(existing_bytes, sheet_name):
    """Returns a set of tuples representing every data row already in the sheet
    (skips the header row), used for duplicate detection on append."""
    wb = load_workbook(io.BytesIO(existing_bytes))
    if sheet_name not in wb.sheetnames:
        return set()
    ws = wb[sheet_name]
    signatures = set()
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue  # header
        signatures.add(tuple("" if v is None else str(v) for v in row))
    return signatures


def append_to_workbook(existing_bytes, header, rows, sheet_name):
    wb = load_workbook(io.BytesIO(existing_bytes))
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        if ws.max_row == 0 or ws.cell(row=1, column=1).value is None:
            ws.append(header)
    else:
        ws = wb.create_sheet(sheet_name)
        ws.append(header)
    for row in rows:
        ws.append(row)
    autosize_columns(ws)
    return wb


def sanitize_filename(name, fallback="extracted_data.xlsx"):
    if not isinstance(name, str) or not name.strip():
        return fallback
    name = name.strip().replace("/", "").replace("\\", "")
    name = re.sub(r"[^A-Za-z0-9 ._-]", "", name)
    name = name.strip(" .")
    if not name:
        return fallback
    if not name.lower().endswith(".xlsx"):
        name += ".xlsx"
    return name[:120]


def sanitize_sheet_name(name, fallback="Extracted"):
    if not isinstance(name, str) or not name.strip():
        return fallback
    name = re.sub(r"[\[\]:*?/\\]", "", name).strip()
    return (name or fallback)[:31]
