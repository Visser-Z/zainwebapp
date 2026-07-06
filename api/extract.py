"""
Vercel Python serverless function: PDF -> preview (step 1 of 2).
Self-contained (no cross-file imports) to avoid Vercel Python bundling issues.

Validates and parses the PDF, returns the extracted header/rows as JSON
so the frontend can show a review table before anything is written.
If an existing spreadsheet is supplied (append mode), also flags how many
of the new rows already appear in that sheet.

The actual .xlsx is only built in /api/build, after the user confirms.
"""

import re
import io
import base64
import logging
import pdfplumber
from openpyxl import load_workbook
from flask import Flask, request, jsonify

app = Flask(__name__)
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger("pdf-extract")

NUMERIC_PATTERN = re.compile(r"^-?\d{1,3}(,\d{3})*(\.\d+)?$|^-?\d+(\.\d+)?$")
MAX_PDF_BYTES = 15 * 1024 * 1024
MAX_EXISTING_XLSX_BYTES = 15 * 1024 * 1024
PDF_MAGIC = b"%PDF-"
XLSX_MAGIC = b"PK"


def to_number_if_possible(value):
    if not isinstance(value, str) or value == "":
        return value
    stripped = value.replace(",", "")
    if NUMERIC_PATTERN.match(value):
        try:
            if "." in stripped:
                return float(stripped)
            return int(stripped)
        except ValueError:
            return value
    return value


def _row_matches_header(clean_row, header_without_page):
    if len(clean_row) != len(header_without_page):
        return False
    return [c.strip().lower() for c in clean_row] == [h.strip().lower() for h in header_without_page]


def extract_rows_from_bytes(pdf_bytes):
    all_rows = []
    header = None
    header_without_page = None
    any_tables_found = False

    with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
        for page_num, page in enumerate(pdf.pages, start=1):
            tables = page.extract_tables()

            if tables:
                any_tables_found = True
                for table in tables:
                    if not table:
                        continue
                    start_idx = 0
                    if header is None:
                        first_row = ["" if c is None else str(c).strip() for c in table[0]]
                        header = ["page"] + first_row
                        header_without_page = first_row
                        start_idx = 1

                    for row in table[start_idx:]:
                        clean_row = ["" if cell is None else str(cell).strip() for cell in row]
                        if header_without_page and _row_matches_header(clean_row, header_without_page):
                            continue
                        typed_row = [to_number_if_possible(v) for v in clean_row]
                        all_rows.append([page_num] + typed_row)
            else:
                text = page.extract_text() or ""
                if not text.strip():
                    continue
                if header is None:
                    header = ["page", "text_line"]
                for line in text.split("\n"):
                    if line.strip():
                        all_rows.append([page_num, line.strip()])

    return header, all_rows, any_tables_found


def get_existing_row_signatures(existing_bytes, sheet_name):
    wb = load_workbook(io.BytesIO(existing_bytes))
    if sheet_name not in wb.sheetnames:
        return set()
    ws = wb[sheet_name]
    signatures = set()
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue
        signatures.add(tuple("" if v is None else str(v) for v in row))
    return signatures


@app.route("/api/extract", methods=["POST"])
def extract():
    data = request.get_json(silent=True)
    if not data:
        return jsonify({"error": "Invalid or missing request body."}), 400

    pdf_b64 = data.get("pdf_base64")
    if not pdf_b64:
        return jsonify({"error": "No PDF provided."}), 400

    try:
        pdf_bytes = base64.b64decode(pdf_b64, validate=True)
    except Exception:
        return jsonify({"error": "The PDF data was corrupted in transit. Please try again."}), 400

    if len(pdf_bytes) == 0:
        return jsonify({"error": "The uploaded PDF is empty."}), 400
    if len(pdf_bytes) > MAX_PDF_BYTES:
        return jsonify({"error": f"PDF is too large (max {MAX_PDF_BYTES // (1024 * 1024)} MB)."}), 400
    if not pdf_bytes.startswith(PDF_MAGIC):
        return jsonify({"error": "That file doesn't look like a valid PDF."}), 400

    try:
        header, rows, used_real_tables = extract_rows_from_bytes(pdf_bytes)
    except Exception:
        logger.exception("PDF parsing failed")
        return jsonify({
            "error": "Couldn't read this PDF. It may be corrupted, password-protected, or a scanned image without selectable text."
        }), 400

    if header is None:
        return jsonify({"error": "No extractable content found in this PDF."}), 400

    duplicate_count = 0
    existing_b64 = data.get("existing_xlsx_base64")
    sheet_name = data.get("sheet_name") or "Extracted"

    if existing_b64:
        try:
            existing_bytes = base64.b64decode(existing_b64, validate=True)
        except Exception:
            return jsonify({"error": "The existing spreadsheet data was corrupted in transit."}), 400

        if len(existing_bytes) > MAX_EXISTING_XLSX_BYTES:
            return jsonify({"error": "The existing spreadsheet is too large."}), 400
        if not existing_bytes.startswith(XLSX_MAGIC):
            return jsonify({"error": "That existing file doesn't look like a valid .xlsx spreadsheet."}), 400

        try:
            existing_signatures = get_existing_row_signatures(existing_bytes, sheet_name)
        except Exception:
            return jsonify({"error": "Couldn't read the existing spreadsheet. It may be corrupted."}), 400

        for row in rows:
            sig = tuple(str(v) for v in row)
            if sig in existing_signatures:
                duplicate_count += 1

    logger.info(
        "Preview: %s rows extracted, %s look like duplicates, real_tables=%s",
        len(rows), duplicate_count, used_real_tables,
    )

    return jsonify({
        "header": header,
        "rows": rows,
        "row_count": len(rows),
        "used_real_tables": used_real_tables,
        "duplicate_count": duplicate_count,
    })
